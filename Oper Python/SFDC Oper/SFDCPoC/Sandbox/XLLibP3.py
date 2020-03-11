from openpyxl import *
import os
import time
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.compat import range

class XLLibp3():
    def __init__(self):
        pass

    def create_workbook(self,file, sheetname):
        sheetexists = 0
        if os.path.isfile(file):

            mywb = load_workbook(file)
            print("XL file exists")
        else:
            print("XL file does not exists")
            mywb = Workbook()
            time.sleep(5)

        sheet = mywb.active

        sheet.title = sheetname
        mywb.save(file)
        time.sleep(5)

    def create_worksheets(self,file, sheetnames):
        sheetexists = 0
        i = 0
        if os.path.isfile(file):
            mywb = load_workbook(file)
            print("XL file exists")
            for sheetname in sheetnames:
                mywb.create_sheet(index=i, title=sheetname)
                mywb.save(file)
                time.sleep(3)


        else:
            print("XL file does not exists")
            mywb = Workbook()
            time.sleep(10)

    def print_XL_cell_values(self,file, sheetname, irow, icolumn):


        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        for row in range(1, mysheet.max_row):
            for col in range(1, mysheet.max_column):
                if mysheet.cell(column=col, row=row).value != 'None':
                    print(mysheet.cell(column=col, row=row).value)

    def get_xl_cell_values(self, file, sheetname, irow, icolumn):
        cellvalues = []
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        for row in range(1, mysheet.max_row):
            for col in range(1, mysheet.max_column):
                if mysheet.cell(column=col, row=row).value != 'None':
                    print(mysheet.cell(column=col, row=row).value)
                    cellvalues.append(mysheet.cell(column=col, row=row).value)
        return cellvalues

    def get_xl_column_letter(self,file, sheetname, columnvalue):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        for col in range(1, mysheet.max_column):
            if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
                id = format(get_column_letter(col))
                break
        return id

    def get_xl_column_index(self,file, sheetname, columnvalue):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        print("max column is " + str(mysheet.max_column))
        for col in range(1, mysheet.max_column):
            # print "***************"
            # print str((mysheet.cell(column=col, row=1).value)).upper()
            # print (str(columnvalue)).upper()
            # print (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper()

            if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
                id = col
                break
        return id

    def get_xl_cell_value(self,file, sheetname, irow, icolumn):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        id = mysheet.cell(column=icolumn, row=irow).value
        return id

    def get_xl_cell_value_using_column_header(self,file, sheetname, irow, columnheader):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        icolumn = XLLibp3.get_xl_column_index(XLLibp3, file, sheetname, columnheader)
        print("The Column is : " + str(icolumn))
        id = mysheet.cell(column=int(icolumn), row=irow).value
        id = str(id).strip()
        return id

    def set_xl_cell_value(self,file, sheetname, irow, icolumn, cellvalue):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        mysheet.cell(column=icolumn, row=irow).value = cellvalue
        mywb.save(file)
        time.sleep(5)

    def set_xl_cell_value_using_column_header(self,file, sheetname, irow, columnheader, cellvalue):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        icolumn = XLLibp3.get_XL_column_index(file, sheetname, columnheader)
        mysheet.cell(column=int(icolumn), row=irow).value = cellvalue
        mywb.save(file)
        time.sleep(5)

    def create_xl_header_lists(self,file, sheetname, lists):
        col = 1
        XLLibp3.Create_Workbook(file, sheetname)
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        for list in lists:
            mysheet.cell(column=int(col), row=1).value = list
            col = int(col) + int(1)
        mywb.save(file)
        time.sleep(5)

    def add_xl_header_column(self,file, sheetname, columnheader):

        XLLibp3.Create_Workbook(file, sheetname)
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        col = int(mysheet.max_column) + int(1)
        mysheet.cell(column=int(col), row=1).value = columnheader
        mywb.save(file)
        time.sleep(5)

    def copy_xl_workbook(self,sourcefile, targetfile):

        if os.path.isfile(sourcefile):
            sourcewb = load_workbook(sourcefile)
            wslists = sourcewb.sheetnames

            for wslist in wslists:
                print("sheet name is : " + str(wslist))
                if os.path.isfile(targetfile):
                    mywb = load_workbook(targetfile)
                    print("XL file exists")
                else:
                    print("XL file does not exists")
                    mywb = Workbook(targetfile)
                    # mywb.save(targetfile)
                    # time.sleep(5)

                sourcews = sourcewb.get_sheet_by_name(str(wslist))
                myws = sourcewb.create_sheet(index=int(sourcewb.get_index(sourcews)), title=str(wslist))
                #   Myws = mywb.active
                myws = sourcewb.copy_worksheet(sourcews)

                mywb.save(targetfile)
                time.sleep(5)

    def get_xl_row_count(self,file, sheetname):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        id = mysheet.max_row
        return id

    def get_XL_column_count(self,file, sheetname):
        id = '0'
        mywb = load_workbook(file)
        mysheet = mywb[sheetname]
        id = mysheet.max_column

        return id

