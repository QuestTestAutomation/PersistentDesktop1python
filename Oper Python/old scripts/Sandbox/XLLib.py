import openpyxl
import os
import time
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.compat import range


def Create_Workbook(file,sheetname):
    sheetexists = 0
    if os.path.isfile(file):
        mywb = openpyxl.load_workbook(file)
        print "XL file exists"
    else:
        print "XL file does not exists"
        mywb = openpyxl.Workbook()
        time.sleep(5)

    sheet = mywb.active

    sheet.title = sheetname
    mywb.save(file)
    time.sleep(5)

def Create_Worksheets(file,sheetnames):
    sheetexists = 0
    i = 0
    if os.path.isfile(file):
        mywb = openpyxl.load_workbook(file)
        print "XL file exists"
        for sheetname in sheetnames:
            mywb.create_sheet(index= i, title= sheetname)
            mywb.save(file)
            time.sleep(3)


    else:
        print "XL file does not exists"
        mywb = openpyxl.Workbook()
        time.sleep(10)



def print_XL_cell_values(file,sheetname,irow,icolumn):
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    for row in range(1, mysheet.max_row):
        for col in range(1, mysheet.max_column):
            if mysheet.cell(column=col, row=row).value <> 'None':
             print mysheet.cell(column=col, row=row).value



def get_XL_column_letter(file,sheetname,columnvalue):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    for col in range(1, mysheet.max_column):
        if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
            id = format(get_column_letter(col))
            break
    return id

def get_XL_column_index(file,sheetname,columnvalue):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    print  "max column is "  + str(mysheet.max_column)
    for col in range(1, mysheet.max_column):
        # print "***************"
        # print str((mysheet.cell(column=col, row=1).value)).upper()
        # print (str(columnvalue)).upper()
        # print (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper()

        if (str(mysheet.cell(column=col, row=1).value)).upper() == (str(columnvalue)).upper():
            id = col
            break
    return id

def get_XL_cell_value(file,sheetname,irow,icolumn):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    id = mysheet.cell(column=icolumn, row=irow).value
    return id

def get_XL_cell_value_using_column_header(file,sheetname,irow,columnheader):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    icolumn = get_XL_column_index(file,sheetname,columnheader)
    print "The Column is : " + str(icolumn)
    id = mysheet.cell(column=int(icolumn), row=irow).value
    id = str(id).strip()
    return id


def set_XL_cell_value(file,sheetname,irow,icolumn,cellvalue):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    mysheet.cell(column=icolumn, row=irow).value = cellvalue
    mywb.save(file)
    time.sleep(5)

def set_XL_cell_value_using_column_header(file,sheetname,irow,columnheader,cellvalue):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    icolumn = get_XL_column_index(file,sheetname,columnheader)
    mysheet.cell(column=int(icolumn), row=irow).value = cellvalue
    mywb.save(file)
    time.sleep(5)

def create_XL_header_lists(file,sheetname,lists):
    col = 1
    Create_Workbook(file, sheetname)
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    for list in lists :
        mysheet.cell(column=int(col), row=1).value = list
        col = int(col) + int(1)
    mywb.save(file)
    time.sleep(5)

def add_XL_header_column(file,sheetname,columnheader):

    Create_Workbook(file, sheetname)
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    col = int(mysheet.max_column) + int(1)
    mysheet.cell(column=int(col), row=1).value = columnheader
    mywb.save(file)
    time.sleep(5)

def copy_XL_workbook(sourcefile,targetfile):


    if os.path.isfile(sourcefile):
        sourcewb = openpyxl.load_workbook(sourcefile)
        wslists = sourcewb.sheetnames

        for wslist in wslists:
            print "sheet name is : " + str(wslist)
            if os.path.isfile(targetfile):
               mywb = openpyxl.load_workbook(targetfile)
               print "XL file exists"
            else:
               print "XL file does not exists"
               mywb = openpyxl.Workbook(targetfile)
               # mywb.save(targetfile)
               # time.sleep(5)

            sourcews = sourcewb.get_sheet_by_name(str(wslist))
            myws = sourcewb.create_sheet(index=int(sourcewb.get_index(sourcews)), title= str(wslist))
            #   Myws = mywb.active
            myws = sourcewb.copy_worksheet(sourcews)

            mywb.save(targetfile)
            time.sleep(5)


def get_XL_row_count(file,sheetname):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    id = mysheet.max_row
    return id

def get_XL_column_count(file,sheetname):
    id = '0'
    mywb = openpyxl.load_workbook(file)
    mysheet = mywb.get_sheet_by_name(sheetname)
    id = mysheet.max_column

    return id
